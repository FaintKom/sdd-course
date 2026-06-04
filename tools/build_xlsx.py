# -*- coding: utf-8 -*-
# Канон 8 недель + гибрид: менеджер отделов в W1, дизайн-система = nice-to-have.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"
INK = "1C1917"; ACCENT = "B5523A"; HEADBG = "1C1917"; BAND = "F0DDD6"; SOFT = "F5F1EA"; OLIVE = "5F6B45"
thin = Side(style="thin", color="D9D2C7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hcell(c, text, bold=True, color="FFFFFF", bg=HEADBG, size=10, wrap=True, align="left"):
    c.value = text
    c.font = Font(name=FONT, bold=bold, color=color, size=size)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    c.border = border

def cell(c, text, bold=False, color=INK, bg=None, size=10, italic=False):
    c.value = text
    c.font = Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = border

wb = Workbook()

ws = wb.active
ws.title = "Программа"
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:H1")
t = ws["A1"]; t.value = "Управляемая разработка с ИИ — программа курса (spec-driven, 8 недель, канон + гибрид)"
t.font = Font(name=FONT, bold=True, size=14, color=ACCENT)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 28
heads = ["Неделя","Модуль","Что делает студент","Артефакт","Боль (опрос)","Материалы","Формат / практика","4C/ID"]
for j,h in enumerate(heads,1): hcell(ws.cell(3,j), h)
prog = [
 ["W1","Spec-Driven Development: почему промпты ломаются","Классифицирует сбой агента; строит ментальную модель «менеджер ↔ отделы» (что отдать ИИ, что проверять); понимает, зачем spec-driven","Карта доверия/контроля + тема проекта","«всё хаотично, нет системы»","4 типа сбоя (контекст/доверие/хаос/качество); менеджер отделов; power inversion","Воркшоп: разбор кейсов диалога; старт сквозного проекта","SI"],
 ["W2","От задачи к спецификации","Пишет спеку (требования, ограничения, EARS-AC, DoD, тесты); clarification gate; прогон агента с/без","Спецификация (5 секций, EARS)","«сносит рабочее решение»","Анатомия спеки; EARS; clarification; примеры хорошо/плохо","Воркшоп: спека на фичу, сравнение с/без","PI · LT"],
 ["W3","Управление контекстом проекта","Формирует контекст-файл (архитектура, конвенции, границы изменений)","AGENTS.md / CLAUDE.md","«не знает контекст проекта»","Constitution/steering; что давать агенту, что скрывать","Воркшоп: контекст-файл + прогон задачи","PI"],
 ["W4","Постановка задач AI-агенту","Ставит задачу в scope, ведёт агента по плану с контрольными точками","Agent plan + PR","«лишние правки, выход за scope»","План для агента; agent mode; quick vs gated","Воркшоп: провести агента по плану","LT"],
 ["W5","Контроль качества AI-кода","Ревьюит AI-код по чек-листу на подсаженных дефектах","Чек-лист ревью + PR","«галлюцинирует, сложно ревьюить»","Чек-лист ревью; типовые дефекты агента","Воркшоп: охота на подсаженные дефекты","PP · PI"],
 ["W6","Тестирование AI-разработки","Пишет тесты из EARS, негативы, использует coverage как критерий, CI","Тесты + coverage, зелёный CI","«доверять нельзя без проверки»","Tests-as-spec; test pyramid; CI","Воркшоп: EARS → тесты + coverage","PP"],
 ["W7","Легаси и существующий код","Вносит фичу/багфикс/рефакторинг через спеку в чужой код без сноса смежного","PR в легаси-репо","«как подключать ИИ к легаси»","Ретрофит SDD на легаси; предотвращение тех-долга","Воркшоп: фича / багфикс в легаси-репо","LT"],
 ["W8","Сквозной проект и защита","Собирает полный цикл, защищает; показывает, как управлял ИИ","Репозиторий + защита","«нет доказательства навыка»","Полный цикл; рубрика защиты; peer review","Защита финального проекта","LT"],
]
r=4
for row in prog:
    for j,v in enumerate(row,1):
        cell(ws.cell(r,j), v, bold=(j in (1,2)))
        if j==5: ws.cell(r,j).font = Font(name=FONT, size=10, italic=True, color=ACCENT)
        if j==8: ws.cell(r,j).font = Font(name=FONT, size=9, bold=True, color=ACCENT)
        if j==1: ws.cell(r,j).fill = PatternFill("solid", fgColor=SOFT)
    ws.row_dimensions[r].height = 84
    r+=1
ws.merge_cells(f"A{r}:H{r}")
n=ws.cell(r,1); n.value="Сквозной проект идёт через все 8 недель (повтор за экспертом на воркшопах). Финал = полный цикл: задача → спека → контекст → реализация агентом → тесты → ревью → доработка → защита."
n.font=Font(name=FONT, size=9, italic=True, color=INK); n.fill=PatternFill("solid",fgColor=BAND)
n.alignment=Alignment(wrap_text=True, vertical="center"); ws.row_dimensions[r].height=42
for c in ws[r]: c.border=border
r+=1
ws.merge_cells(f"A{r}:H{r}")
o=ws.cell(r,1); o.value="Опционально (nice-to-have, вне канона): дизайн-система как спека для UI (frontend/fullstack, врезка W3–W4). Enrichment/v2: безопасность AI-разработки; локальные модели/стоимость (+ бесплатный путь); governance на команду."
o.font=Font(name=FONT, size=9, italic=True, color=OLIVE); o.fill=PatternFill("solid",fgColor="F4F1E8")
o.alignment=Alignment(wrap_text=True, vertical="center"); ws.row_dimensions[r].height=42
for c in ws[r]: c.border=border
widths=[8,30,34,22,24,30,30,9]
for j,w in enumerate(widths,1): ws.column_dimensions[chr(64+j)].width=w
ws.freeze_panes="A4"

ws2 = wb.create_sheet("Образовательные результаты")
ws2.sheet_view.showGridLines=False
ws2.merge_cells("A1:G1")
t=ws2["A1"]; t.value="Образовательные результаты в формате ABCD (Audience · Behavior · Condition · Degree) + 4C/ID"
t.font=Font(name=FONT,bold=True,size=13,color=ACCENT); t.alignment=Alignment(vertical="center")
ws2.row_dimensions[1].height=26
heads2=["ID","Уровень","A — кто","B — что делает (наблюдаемый глагол)","C — при каких условиях","D — критерий / степень","4C/ID"]
for j,h in enumerate(heads2,1): hcell(ws2.cell(3,j),h)
def section(ws,r,title):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
    c=ws.cell(r,1); c.value=title
    c.font=Font(name=FONT,bold=True,size=10,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=ACCENT)
    c.alignment=Alignment(vertical="center")
    for cc in ws[r]: cc.border=border
    ws.row_dimensions[r].height=20
    return r+1
prog_or=[
 ["ОР-1","Программа","Выпускник курса","проводит задачу (фича/багфикс/рефактор) через полный spec-driven цикл, управляя AI-агентом","в репозитории, на защите перед экспертом","доходит до всех критериев готовности без ручного переписывания сгенерённого кода","LT"],
 ["ОР-2","Программа","Выпускник","формализует задачу в спецификацию (требования, ограничения, EARS-AC, DoD, тесты)","для реальной/учебной задачи по шаблону","спека покрывает 5 секций и однозначно ограничивает агента","PI"],
 ["ОР-3","Программа","Выпускник","управляет контекстом проекта (constitution/AGENTS.md/CLAUDE.md)","в репозитории","агент выполняет задачу не выходя за границы и не трогая запрещённые файлы","PI"],
 ["ОР-4","Программа","Выпускник","ставит задачу агенту в scope и ведёт по плану с контрольными точками","на ограниченном scope в репозитории","получает diff/PR, соответствующий спеке; лишние изменения отклонены","LT"],
 ["ОР-5","Программа","Выпускник","проверяет AI-generated code через тесты и чек-лист ревью","имея diff и спеку (в т.ч. с дефектами)","выявляет ≥80% типовых дефектов до мерджа","PP"],
 ["ОР-6","Программа","Выпускник","применяет ИИ в легаси и встраивает spec-driven в личный/командный процесс","в легаси-репозитории, не им написанном","изменение проходит тесты, не ломает смежное, ограничено scope; процесс повторяем","LT"],
]
mod_or=[
 ["М1","W1 · Почему промпты / менеджер отделов","студент","классифицирует сбой агента и распределяет задачи по «отделам» (доверие vs контроль)","на 4 кейсах диалога с ИИ","верно атрибутирует причину сбоя в ≥3 из 4","SI"],
 ["М2","W2 · Спецификация","студент","составляет спеку (требования, ограничения, EARS-AC, DoD, тесты), помечает NEEDS CLARIFICATION","для задачи своего проекта","5 секций покрыты; результат с/без спеки расходится","PI · LT"],
 ["М3","W3 · Контекст","студент","формирует контекст-файл (архитектура, конвенции, границы)","для своего/учебного репо","агент не выходит за границы, не трогает запретное","PI"],
 ["М4","W4 · Постановка задач","студент","ставит задачу в scope и ведёт агента по плану с контрольными точками","на атомарной задаче в репо","diff соответствует спеке, лишнее отклонено","LT"],
 ["М5","W5 · Контроль качества","студент","ревьюит AI-код по чек-листу","на diff с подсаженными дефектами","находит ≥80% типовых дефектов до мерджа","PP · PI"],
 ["М6","W6 · Тестирование","студент","проектирует тест-стратегию из EARS и использует coverage как критерий","для AI-сгенерённых изменений","критичная логика покрыта, есть негативы, CI зелёный","PP"],
 ["М7","W7 · Легаси","студент","вносит фичу/багфикс/рефакторинг через спеку в существующий код","в легаси-репо, не им написанном","изменение проходит тесты, не ломает смежное, scope соблюдён","LT"],
 ["М8","W8 · Проект + защита","студент","демонстрирует полный spec-driven цикл и то, как управлял ИИ","на защите, репозиторий открыт","проходит рубрику защиты (6 критериев)","LT"],
]
r=4
r=section(ws2,r,"Уровень программы (terminal objectives)")
for row in prog_or:
    for j,v in enumerate(row,1):
        cell(ws2.cell(r,j),v,bold=(j==1))
        if j==1: ws2.cell(r,j).font=Font(name=FONT,bold=True,color=ACCENT,size=10)
        if j==7: ws2.cell(r,j).font=Font(name=FONT,bold=True,color=ACCENT,size=9)
    ws2.row_dimensions[r].height=64; r+=1
r=section(ws2,r,"Уровень модулей")
for row in mod_or:
    for j,v in enumerate(row,1):
        cell(ws2.cell(r,j),v,bold=(j==1))
        if j==1: ws2.cell(r,j).font=Font(name=FONT,bold=True,color=ACCENT,size=10)
        if j==7: ws2.cell(r,j).font=Font(name=FONT,bold=True,color=ACCENT,size=9)
    ws2.row_dimensions[r].height=58; r+=1
widths2=[7,28,12,40,28,40,9]
for j,w in enumerate(widths2,1): ws2.column_dimensions[chr(64+j)].width=w
ws2.freeze_panes="A4"

ws3=wb.create_sheet("Боли (опрос) → модуль")
ws3.sheet_view.showGridLines=False
ws3.merge_cells("A1:D1")
t=ws3["A1"]; t.value="Боли студентов из опроса → модуль программы (выборка мелкая, не репрезентативная)"
t.font=Font(name=FONT,bold=True,size=13,color=ACCENT); t.alignment=Alignment(vertical="center")
ws3.row_dimensions[1].height=26
for j,h in enumerate(["Боль / цитата","Тип","Модуль","Как закрываем"],1): hcell(ws3.cell(3,j),h)
pains=[
 ["«всё хаотично, нет системного подхода»","closed","W1","Ментальная модель «менеджер отделов» + системный spec-driven воркфлоу"],
 ["«сносит рабочее решение, подгоняет проект под своё»","free-text","W2","Спека (EARS) ограничивает агента рамками задачи"],
 ["«не знает контекст проекта»","closed/free","W3","Контекст-файл проекта (constitution) как вводные для агента"],
 ["«галлюцинирует, нельзя доверять»","closed","W5","Тесты + ревью как контроль качества"],
 ["«код работает, но плохого качества»","closed","W5","Чек-лист ревью AI-кода"],
 ["«сложно ревьюить AI-код»","closed","W5","Чек-лист + тренировка на подсаженных дефектах"],
 ["«с доверием беда»","free-text","W5–W6","Контроль качества: ревью + тесты"],
 ["«работа с легаси»","closed","W7","Ретрофит SDD на существующий код"],
 ["«курсы поверхностные, устаревают, нет структуры»","free-text","вся программа","Упор на метод (воркфлоу), а не на инструменты, которые протухнут"],
 ["«нет структуры, разнобой UI»","free-text","опц · W3–W4","Дизайн-система как спека для UI (nice-to-have, frontend/fullstack)"],
 ["«безопасность: слив данных, лазейки, MCP, ИИ управляет консолью»","free-text","enrichment / v2","Отдельный модуль безопасности AI-разработки (вне канона)"],
 ["«дорогие токены, как построить локальную систему»","free-text","enrichment / pre-work","Локальные модели + задокументированный бесплатный путь"],
]
r=4
for row in pains:
    for j,v in enumerate(row,1):
        cell(ws3.cell(r,j),v)
        if j==1: ws3.cell(r,j).font=Font(name=FONT,italic=True,color=ACCENT,size=10)
        if j==3: ws3.cell(r,j).font=Font(name=FONT,bold=True,color=INK,size=10)
    ws3.row_dimensions[r].height=40; r+=1
for j,w in enumerate([46,12,18,44],1): ws3.column_dimensions[chr(64+j)].width=w
ws3.freeze_panes="A4"

ws4=wb.create_sheet("Профиль ЦА")
ws4.sheet_view.showGridLines=False
ws4.merge_cells("A1:B1")
t=ws4["A1"]; t.value="Профиль ЦА по опросу"
t.font=Font(name=FONT,bold=True,size=13,color=ACCENT); t.alignment=Alignment(vertical="center")
ws4.row_dimensions[1].height=26
for j,h in enumerate(["Параметр","Значения (из опроса)"],1): hcell(ws4.cell(3,j),h)
ca=[
 ["Роли","backend (API, БД), frontend, mobile, devops, fullstack"],
 ["Уровень","junior+ → senior / lead; есть стажёры"],
 ["Стек","JS/TS, Python, Java/Kotlin, Go, C#"],
 ["Частота ИИ","большинство — «каждый день» / «основной инструмент»"],
 ["Инструменты","Copilot, Cursor, Claude Code, ChatGPT, Claude.ai, DeepSeek, Gemini, Гигачат, локальные модели"],
 ["Ускорение","разброс: «ускоряет отдельные задачи» / «сильно влияет» / «почти не ускоряет»"],
 ["Цена","«только бесплатно» → «до 20 000»; много «не могу оценить» → ценообразование на данных опроса не строить"],
 ["Выборка","мелкая, не репрезентативная; есть троллинг-ответы — фильтровать"],
 ["Гейт","по способности пройти петлю: читает код, Git, тесты, ставит задачу, кодит на 1+ языке"],
 ["Тиры","T1 junior+ / T2 middle (ядро) / T3 senior-lead — scaffolding fade"],
]
r=4
for row in ca:
    for j,v in enumerate(row,1):
        cell(ws4.cell(r,j),v,bold=(j==1))
    ws4.row_dimensions[r].height=34; r+=1
ws4.column_dimensions["A"].width=18; ws4.column_dimensions["B"].width=72
ws4.freeze_panes="A4"

import os
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "01-deliverables", "Программа_курса.xlsx")
wb.save(OUT)
print("saved", OUT)
